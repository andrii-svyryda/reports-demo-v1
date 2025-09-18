import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
import string
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis
import warnings
warnings.filterwarnings('ignore')

# Create public folder if it doesn't exist
public_folder = os.path.join(os.path.dirname(__file__), 'public')
os.makedirs(public_folder, exist_ok=True)

# Generate synthetic data for Epic System Integration
def generate_patient_data(num_patients=500):
    """Generate synthetic patient data"""
    patients = []
    for i in range(num_patients):
        mrn = f"MRN{str(i+1000000).zfill(7)}"
        patients.append({
            'MRN': mrn,
            'FirstName': ''.join(random.choices(string.ascii_uppercase, k=8)),
            'LastName': ''.join(random.choices(string.ascii_uppercase, k=10)),
            'DOB': (datetime.now() - timedelta(days=random.randint(365, 30000))).strftime('%Y-%m-%d'),
            'Gender': random.choice(['M', 'F']),
            'Phone': f"{random.randint(100,999)}-{random.randint(100,999)}-{random.randint(1000,9999)}",
            'Email': f"patient{i}@email.com",
            'InsuranceID': f"INS{random.randint(100000,999999)}",
            'TenantID': f"TENANT_{random.randint(1,5):03d}",
            'LastSync': datetime.now() - timedelta(minutes=random.randint(0, 10080))
        })
    return pd.DataFrame(patients)

def generate_lab_orders(num_orders=1500):
    """Generate synthetic laboratory orders"""
    orders = []
    test_types = ['CBC', 'BMP', 'CMP', 'Lipid Panel', 'HbA1c', 'TSH', 'Urinalysis', 'PT/INR', 'Blood Culture', 'COVID-19 PCR']
    priorities = ['STAT', 'URGENT', 'ROUTINE']
    statuses = ['PENDING', 'COLLECTED', 'PROCESSING', 'RESULTED', 'CANCELLED']

    for i in range(num_orders):
        order_date = datetime.now() - timedelta(days=random.randint(0, 90))
        orders.append({
            'OrderID': f"ORD{str(i+2000000).zfill(8)}",
            'MRN': f"MRN{str(random.randint(1000000, 1000499)).zfill(7)}",
            'TestCode': f"TC{random.randint(1000,9999)}",
            'TestName': random.choice(test_types),
            'Priority': random.choice(priorities),
            'Status': random.choice(statuses),
            'OrderDateTime': order_date,
            'CollectionDateTime': order_date + timedelta(hours=random.randint(1, 24)) if random.random() > 0.2 else None,
            'ResultDateTime': order_date + timedelta(hours=random.randint(24, 72)) if random.random() > 0.3 else None,
            'Provider': f"DR_{random.randint(100,999)}",
            'Department': random.choice(['ED', 'ICU', 'Medicine', 'Surgery', 'Pediatrics', 'OB/GYN'])
        })
    return pd.DataFrame(orders)

def generate_specimen_tracking(num_specimens=2000):
    """Generate specimen tracking data with QR codes"""
    specimens = []
    locations = ['Collection Station', 'Transport', 'Lab Reception', 'Processing Area', 'Analyzer', 'Storage', 'Disposal']

    for i in range(num_specimens):
        specimens.append({
            'SpecimenID': f"SPEC{str(i+3000000).zfill(8)}",
            'QRCode': ''.join(random.choices(string.ascii_uppercase + string.digits, k=12)),
            'OrderID': f"ORD{str(random.randint(2000000, 2001499)).zfill(8)}",
            'TubeType': random.choice(['EDTA', 'SST', 'Heparin', 'Citrate', 'Urine Cup']),
            'Volume': round(random.uniform(1.0, 10.0), 1),
            'CollectedBy': f"TECH_{random.randint(100,999)}",
            'CurrentLocation': random.choice(locations),
            'Temperature': round(random.uniform(2.0, 8.0), 1),
            'ChainOfCustody': random.randint(1, 10),
            'Timestamp': datetime.now() - timedelta(hours=random.randint(0, 168))
        })
    return pd.DataFrame(specimens)

def generate_test_results(num_results=3000):
    """Generate test results data"""
    results = []
    result_statuses = ['Normal', 'Abnormal', 'Critical', 'Pending Review']

    for i in range(num_results):
        results.append({
            'ResultID': f"RES{str(i+4000000).zfill(8)}",
            'OrderID': f"ORD{str(random.randint(2000000, 2001499)).zfill(8)}",
            'TestComponent': random.choice(['WBC', 'RBC', 'Hemoglobin', 'Glucose', 'Creatinine', 'Sodium', 'Potassium']),
            'Value': round(random.uniform(0.5, 200.0), 2),
            'Units': random.choice(['mg/dL', 'mmol/L', 'g/dL', '10^9/L', '%']),
            'ReferenceRange': f"{round(random.uniform(0, 50), 1)}-{round(random.uniform(51, 200), 1)}",
            'Status': random.choice(result_statuses),
            'VerifiedBy': f"PATH_{random.randint(100,999)}" if random.random() > 0.3 else None,
            'ResultDateTime': datetime.now() - timedelta(hours=random.randint(0, 72)),
            'CriticalNotified': random.choice([True, False]) if random.choice(result_statuses) == 'Critical' else False
        })
    return pd.DataFrame(results)

def generate_sync_logs(num_logs=5000):
    """Generate synchronization logs"""
    logs = []
    sync_types = ['PATIENT_DEMOGRAPHICS', 'LAB_ORDERS', 'TEST_RESULTS', 'SPECIMEN_STATUS', 'INSURANCE_INFO']
    sync_statuses = ['SUCCESS', 'FAILED', 'PARTIAL', 'RETRY', 'TIMEOUT']
    error_codes = ['NONE', 'AUTH_FAILED', 'NETWORK_ERROR', 'DATA_VALIDATION', 'EPIC_UNAVAILABLE', 'RATE_LIMIT']

    for i in range(num_logs):
        logs.append({
            'LogID': f"LOG{str(i+5000000).zfill(8)}",
            'SyncType': random.choice(sync_types),
            'Direction': random.choice(['EPIC_TO_LIMS', 'LIMS_TO_EPIC']),
            'Status': random.choice(sync_statuses),
            'RecordsProcessed': random.randint(0, 1000),
            'RecordsFailed': random.randint(0, 50),
            'Duration': random.randint(100, 10000),
            'ErrorCode': random.choice(error_codes),
            'Timestamp': datetime.now() - timedelta(minutes=random.randint(0, 10080)),
            'TenantID': f"TENANT_{random.randint(1,5):03d}"
        })
    return pd.DataFrame(logs)

def generate_performance_metrics(num_days=30):
    """Generate daily performance metrics"""
    metrics = []
    base_date = datetime.now() - timedelta(days=num_days)

    for day in range(num_days):
        current_date = base_date + timedelta(days=day)
        metrics.append({
            'Date': current_date.strftime('%Y-%m-%d'),
            'TotalOrders': random.randint(200, 500),
            'CompletedTests': random.randint(180, 450),
            'AverageTAT': round(random.uniform(2.0, 8.0), 2),
            'CriticalValues': random.randint(0, 15),
            'SpecimensCollected': random.randint(300, 600),
            'SyncSuccess': round(random.uniform(0.92, 0.99), 3),
            'SystemUptime': round(random.uniform(0.985, 0.999), 3),
            'APICallsCount': random.randint(5000, 15000),
            'ErrorRate': round(random.uniform(0.001, 0.05), 3)
        })
    return pd.DataFrame(metrics)

# Create raw input Excel file
def create_raw_excel():
    """Create the raw input Excel file with multiple sheets of complex data"""
    filepath = os.path.join(public_folder, 'input-report.xlsx')

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Generate and write all data sheets
        generate_patient_data().to_excel(writer, sheet_name='RAW_PATIENTS', index=False)
        generate_lab_orders().to_excel(writer, sheet_name='RAW_ORDERS', index=False)
        generate_specimen_tracking().to_excel(writer, sheet_name='RAW_SPECIMENS', index=False)
        generate_test_results().to_excel(writer, sheet_name='RAW_RESULTS', index=False)
        generate_sync_logs().to_excel(writer, sheet_name='SYNC_LOGS', index=False)
        generate_performance_metrics().to_excel(writer, sheet_name='PERF_METRICS', index=False)

    print(f"Raw input file created: {filepath}")
    return filepath

# Create human-friendly report
def create_friendly_report():
    """Create the human-friendly Excel report with charts and formatted data"""
    # First, read the raw data
    raw_file = os.path.join(public_folder, 'input-report.xlsx')

    # Read all sheets
    patients_df = pd.read_excel(raw_file, sheet_name='RAW_PATIENTS')
    orders_df = pd.read_excel(raw_file, sheet_name='RAW_ORDERS')
    specimens_df = pd.read_excel(raw_file, sheet_name='RAW_SPECIMENS')
    results_df = pd.read_excel(raw_file, sheet_name='RAW_RESULTS')
    sync_logs_df = pd.read_excel(raw_file, sheet_name='SYNC_LOGS')
    performance_df = pd.read_excel(raw_file, sheet_name='PERF_METRICS')

    # Create workbook
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2E75B6")
    subtitle_font = Font(bold=True, size=14, color="4472C4")
    data_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1['A1'] = "Epic System Integration - Laboratory Management Dashboard"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:F1')

    ws1['A3'] = "Report Generated:"
    ws1['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws1['B3'].font = Font(italic=True)

    # Key Metrics Summary
    ws1['A5'] = "KEY PERFORMANCE INDICATORS"
    ws1['A5'].font = subtitle_font
    ws1.merge_cells('A5:F5')

    # Calculate KPIs
    total_patients = len(patients_df)
    total_orders = len(orders_df)
    completed_tests = len(orders_df[orders_df['Status'] == 'RESULTED'])
    avg_tat = performance_df['AverageTAT'].mean()
    sync_success_rate = sync_logs_df[sync_logs_df['Status'] == 'SUCCESS'].shape[0] / len(sync_logs_df) * 100
    critical_values_total = results_df[results_df['Status'] == 'Critical'].shape[0]

    kpi_data = [
        ['Metric', 'Value', 'Status', 'Target', 'Achievement'],
        ['Total Active Patients', f'{total_patients:,}', 'On Track', '450', f'{(total_patients/450*100):.1f}%'],
        ['Total Lab Orders', f'{total_orders:,}', 'Excellent', '1,200', f'{(total_orders/1200*100):.1f}%'],
        ['Tests Completed', f'{completed_tests:,}', 'Good', '1,000', f'{(completed_tests/1000*100):.1f}%'],
        ['Average TAT (hours)', f'{avg_tat:.2f}', 'Good', '6.0', f'{(6/avg_tat*100):.1f}%'],
        ['Sync Success Rate', f'{sync_success_rate:.1f}%', 'Excellent', '95%', f'{(sync_success_rate/95*100):.1f}%'],
        ['Critical Values Reported', f'{critical_values_total:,}', 'Normal', 'N/A', 'N/A']
    ]

    for row_idx, row_data in enumerate(kpi_data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 7:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            cell.border = data_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Sheet 2: Test Volume Analysis
    ws2 = wb.create_sheet("Test Volume Analysis")
    ws2['A1'] = "Laboratory Test Volume Analysis"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:D1')

    # Group orders by test type
    test_summary = orders_df.groupby('TestName').agg({
        'OrderID': 'count',
        'Priority': lambda x: (x == 'STAT').sum()
    }).reset_index()
    test_summary.columns = ['Test Type', 'Total Orders', 'STAT Orders']
    test_summary['% STAT'] = (test_summary['STAT Orders'] / test_summary['Total Orders'] * 100).round(1)
    test_summary = test_summary.sort_values('Total Orders', ascending=False)

    # Write test summary
    ws2['A3'] = "Test Type Distribution"
    ws2['A3'].font = subtitle_font

    for col_idx, header in enumerate(['Test Type', 'Total Orders', 'STAT Orders', '% STAT'], start=1):
        cell = ws2.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = data_border

    for row_idx, row in test_summary.iterrows():
        for col_idx, value in enumerate(row, start=1):
            cell = ws2.cell(row=row_idx+6, column=col_idx, value=value)
            cell.border = data_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='center')

    # Add bar chart for test volumes
    chart1 = BarChart()
    chart1.title = "Test Volume by Type"
    chart1.x_axis.title = "Test Type"
    chart1.y_axis.title = "Number of Orders"
    chart1.height = 10
    chart1.width = 15

    data = Reference(ws2, min_col=2, min_row=5, max_row=5+len(test_summary), max_col=2)
    categories = Reference(ws2, min_col=1, min_row=6, max_row=5+len(test_summary))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(categories)
    ws2.add_chart(chart1, "F5")

    # Sheet 3: TAT Performance
    ws3 = wb.create_sheet("TAT Performance")
    ws3['A1'] = "Turnaround Time Performance"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:E1')

    # Department TAT analysis
    dept_tat = orders_df.groupby('Department').agg({
        'OrderID': 'count'
    }).reset_index()
    dept_tat.columns = ['Department', 'Total Orders']
    dept_tat['Avg TAT (hrs)'] = [round(random.uniform(2, 8), 2) for _ in range(len(dept_tat))]
    dept_tat['Within Target'] = [f"{random.randint(85, 99)}%" for _ in range(len(dept_tat))]
    dept_tat = dept_tat.sort_values('Total Orders', ascending=False)

    ws3['A3'] = "Department-wise TAT Analysis"
    ws3['A3'].font = subtitle_font

    for col_idx, header in enumerate(['Department', 'Total Orders', 'Avg TAT (hrs)', 'Within Target'], start=1):
        cell = ws3.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = data_border

    for row_idx, row in dept_tat.iterrows():
        for col_idx, value in enumerate(row, start=1):
            cell = ws3.cell(row=row_idx+6, column=col_idx, value=value)
            cell.border = data_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal='center')

    # Add line chart for TAT trend
    chart2 = LineChart()
    chart2.title = "Daily Average TAT Trend"
    chart2.x_axis.title = "Date"
    chart2.y_axis.title = "TAT (hours)"
    chart2.height = 10
    chart2.width = 15

    # Write performance metrics for chart
    perf_start_row = 15
    ws3['A14'] = "Daily TAT Trend"
    ws3['A14'].font = subtitle_font

    for col_idx, header in enumerate(['Date', 'Avg TAT'], start=1):
        cell = ws3.cell(row=perf_start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in performance_df[['Date', 'AverageTAT']].iterrows():
        ws3.cell(row=row_idx+perf_start_row+1, column=1, value=row['Date'])
        ws3.cell(row=row_idx+perf_start_row+1, column=2, value=row['AverageTAT'])

    data = Reference(ws3, min_col=2, min_row=perf_start_row, max_row=perf_start_row+len(performance_df), max_col=2)
    dates = Reference(ws3, min_col=1, min_row=perf_start_row+1, max_row=perf_start_row+len(performance_df))
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(dates)
    ws3.add_chart(chart2, "F15")

    # Sheet 4: System Integration Status
    ws4 = wb.create_sheet("Integration Status")
    ws4['A1'] = "Epic-LIMS Integration Status"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:E1')

    # Sync status summary
    sync_summary = sync_logs_df.groupby(['SyncType', 'Status']).size().unstack(fill_value=0)

    ws4['A3'] = "Synchronization Performance by Type"
    ws4['A3'].font = subtitle_font

    # Write sync summary
    start_row = 5
    ws4.cell(row=start_row, column=1, value="Sync Type").font = header_font
    ws4.cell(row=start_row, column=1).fill = header_fill

    for col_idx, status in enumerate(sync_summary.columns, start=2):
        cell = ws4.cell(row=start_row, column=col_idx, value=status)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, (sync_type, row_data) in enumerate(sync_summary.iterrows(), start=start_row+1):
        ws4.cell(row=row_idx, column=1, value=sync_type)
        for col_idx, value in enumerate(row_data, start=2):
            ws4.cell(row=row_idx, column=col_idx, value=value)

    # Add pie chart for sync status
    chart3 = PieChart()
    chart3.title = "Overall Sync Status Distribution"
    chart3.height = 10
    chart3.width = 10

    sync_status_counts = sync_logs_df['Status'].value_counts()

    ws4['A15'] = "Overall Sync Status"
    ws4['A15'].font = subtitle_font

    for idx, (status, count) in enumerate(sync_status_counts.items(), start=17):
        ws4.cell(row=idx, column=1, value=status)
        ws4.cell(row=idx, column=2, value=count)

    data = Reference(ws4, min_col=2, min_row=17, max_row=16+len(sync_status_counts))
    labels = Reference(ws4, min_col=1, min_row=17, max_row=16+len(sync_status_counts))
    chart3.add_data(data)
    chart3.set_categories(labels)
    ws4.add_chart(chart3, "D15")

    # Sheet 5: Specimen Tracking
    ws5 = wb.create_sheet("Specimen Tracking")
    ws5['A1'] = "Specimen Chain of Custody Analysis"
    ws5['A1'].font = title_font
    ws5.merge_cells('A1:D1')

    # Location distribution
    location_summary = specimens_df['CurrentLocation'].value_counts().reset_index()
    location_summary.columns = ['Location', 'Count']
    location_summary['Percentage'] = (location_summary['Count'] / location_summary['Count'].sum() * 100).round(1)

    ws5['A3'] = "Current Specimen Locations"
    ws5['A3'].font = subtitle_font

    for col_idx, header in enumerate(['Location', 'Count', 'Percentage'], start=1):
        cell = ws5.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in location_summary.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws5.cell(row=row_idx+6, column=col_idx, value=value)

    # Sheet 6: Multi-Tenant Analytics
    ws6 = wb.create_sheet("Multi-Tenant Analytics")
    ws6['A1'] = "Multi-Tenant System Usage"
    ws6['A1'].font = title_font
    ws6.merge_cells('A1:D1')

    # Tenant usage summary
    tenant_summary = patients_df['TenantID'].value_counts().reset_index()
    tenant_summary.columns = ['Tenant', 'Patient Count']

    # Add more tenant metrics
    tenant_orders = sync_logs_df.groupby('TenantID')['RecordsProcessed'].sum().reset_index()
    tenant_orders.columns = ['Tenant', 'Records Processed']

    tenant_summary = tenant_summary.merge(tenant_orders, on='Tenant', how='left')
    tenant_summary['Avg Records/Patient'] = (tenant_summary['Records Processed'] / tenant_summary['Patient Count']).round(1)

    ws6['A3'] = "Tenant Usage Statistics"
    ws6['A3'].font = subtitle_font

    for col_idx, header in enumerate(['Tenant', 'Patient Count', 'Records Processed', 'Avg Records/Patient'], start=1):
        cell = ws6.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in tenant_summary.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws6.cell(row=row_idx+6, column=col_idx, value=value)

    # Adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
            except AttributeError:
                continue
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save the report
    filepath = os.path.join(public_folder, 'sample-report.xlsx')
    wb.save(filepath)
    print(f"Human-friendly report created: {filepath}")
    return filepath

# Main execution
if __name__ == "__main__":
    print("Epic System Integration - Report Generator")
    print("==========================================")
    print()

    print("Step 1: Creating raw input Excel file with complex data...")
    raw_file = create_raw_excel()
    print(f"[OK] Raw data file created with 6 sheets containing {5000}+ records")
    print()

    print("Step 2: Processing data and creating human-friendly report...")
    report_file = create_friendly_report()
    print(f"[OK] Human-friendly report created with:")
    print("  - Executive Summary with KPIs")
    print("  - Test Volume Analysis with bar charts")
    print("  - TAT Performance with trend charts")
    print("  - Integration Status with pie charts")
    print("  - Specimen Tracking analytics")
    print("  - Multi-Tenant usage statistics")
    print()

    print("Report generation complete!")
    print(f"Input file: {raw_file}")
    print(f"Output file: {report_file}")