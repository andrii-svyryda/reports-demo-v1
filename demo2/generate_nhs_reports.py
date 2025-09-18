import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
import os
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis
import string

script_folder = os.path.dirname(os.path.abspath(__file__))
public_folder = os.path.join(script_folder, 'public')
os.makedirs(public_folder, exist_ok=True)

def generate_patient_ids(count):
    """Generate NHS-style patient IDs"""
    return [f"NHS{random.randint(1000000000, 9999999999)}" for _ in range(count)]

def generate_snomed_codes():
    """Generate realistic SNOMED CT codes for conditions"""
    conditions = {
        '73211009': 'Diabetes mellitus',
        '38341003': 'Hypertension',
        '195967001': 'Asthma',
        '13645005': 'COPD',
        '53741008': 'Coronary heart disease',
        '84114007': 'Heart failure',
        '396275006': 'Osteoarthritis',
        '35489007': 'Depression',
        '197480006': 'Anxiety disorder',
        '49436004': 'Atrial fibrillation'
    }
    return conditions

def generate_medication_codes():
    """Generate DM&D medication codes"""
    medications = {
        '318185001': 'Metformin 500mg tablets',
        '319773006': 'Amlodipine 5mg tablets',
        '376584008': 'Salbutamol 100mcg inhaler',
        '374804007': 'Simvastatin 40mg tablets',
        '391761004': 'Ramipril 5mg capsules',
        '322236009': 'Paracetamol 500mg tablets',
        '387517004': 'Omeprazole 20mg capsules',
        '387458008': 'Aspirin 75mg tablets',
        '386845007': 'Levothyroxine 100mcg tablets',
        '387525002': 'Furosemide 40mg tablets'
    }
    return medications

def generate_raw_data():
    """Generate raw NHS data for input file"""
    np.random.seed(42)
    random.seed(42)

    # Generate patient demographics
    num_patients = 500
    patient_ids = generate_patient_ids(num_patients)

    demographics_data = {
        'patient_id': patient_ids,
        'nhs_number': [random.randint(1000000000, 9999999999) for _ in range(num_patients)],
        'dob': [datetime(1940, 1, 1) + timedelta(days=random.randint(0, 30000)) for _ in range(num_patients)],
        'gender_code': np.random.choice([1, 2, 9], num_patients, p=[0.48, 0.48, 0.04]),
        'ethnicity_code': np.random.choice(range(1, 19), num_patients),
        'gp_practice_code': [''.join(random.choices(string.ascii_uppercase + string.digits, k=6)) for _ in range(num_patients)],
        'lsoa_code': ['E0' + str(random.randint(1000000, 9999999)) for _ in range(num_patients)],
        'imd_decile': np.random.randint(1, 11, num_patients)
    }

    # Generate diagnoses data
    conditions = generate_snomed_codes()
    diagnoses_records = []
    for patient_id in patient_ids[:300]:  # Not all patients have diagnoses
        num_conditions = np.random.poisson(2) + 1
        for _ in range(num_conditions):
            diagnoses_records.append({
                'patient_id': patient_id,
                'snomed_code': random.choice(list(conditions.keys())),
                'diagnosis_date': datetime.now() - timedelta(days=random.randint(0, 1825)),
                'status_code': np.random.choice([1, 2, 3], p=[0.7, 0.2, 0.1]),  # Active, Resolved, Inactive
                'severity_score': np.random.uniform(0.1, 10.0),
                'confidence_level': np.random.uniform(0.6, 1.0)
            })

    # Generate medications data
    meds = generate_medication_codes()
    medications_records = []
    for patient_id in patient_ids[:350]:
        num_meds = np.random.poisson(3) + 1
        for _ in range(num_meds):
            medications_records.append({
                'patient_id': patient_id,
                'dm_d_code': random.choice(list(meds.keys())),
                'start_date': datetime.now() - timedelta(days=random.randint(0, 730)),
                'daily_dose': np.random.choice([1, 2, 3, 4]),
                'quantity': np.random.randint(28, 84),
                'status': np.random.choice([1, 2, 3]),  # Active, Discontinued, On-hold
                'adherence_score': np.random.uniform(0.3, 1.0)
            })

    # Generate appointments data
    appointments_records = []
    for _ in range(2000):
        appointments_records.append({
            'patient_id': random.choice(patient_ids),
            'appointment_date': datetime.now() + timedelta(days=random.randint(-365, 90)),
            'specialty_code': np.random.choice(range(100, 900)),
            'appointment_type': np.random.choice([1, 2, 3, 4]),  # New, Follow-up, Emergency, Telephone
            'status': np.random.choice([1, 2, 3, 4, 5]),  # Scheduled, Completed, Cancelled, No-show, Rescheduled
            'wait_time_days': np.random.randint(0, 180),
            'consultation_duration': np.random.randint(5, 60)
        })

    # Generate test results data
    test_results = []
    test_types = ['HBA1C', 'CHOL', 'BP_SYS', 'BP_DIA', 'BMI', 'EGFR', 'CRP', 'TSH', 'B12', 'VITD']
    for _ in range(3000):
        test_results.append({
            'patient_id': random.choice(patient_ids),
            'test_code': random.choice(test_types),
            'result_value': np.random.uniform(0.5, 200),
            'test_date': datetime.now() - timedelta(days=random.randint(0, 365)),
            'abnormal_flag': np.random.choice([0, 1, 2]),  # Normal, High, Low
            'reference_min': np.random.uniform(0, 50),
            'reference_max': np.random.uniform(50, 200),
            'unit_code': np.random.choice([1, 2, 3, 4, 5])
        })

    # Generate admission data
    admissions = []
    for _ in range(500):
        admission_date = datetime.now() - timedelta(days=random.randint(0, 730))
        los = random.randint(1, 30)
        admissions.append({
            'patient_id': random.choice(patient_ids),
            'admission_date': admission_date,
            'discharge_date': admission_date + timedelta(days=los),
            'ward_code': ''.join(random.choices(string.ascii_uppercase, k=3)) + str(random.randint(1, 9)),
            'admission_method': np.random.choice(range(11, 31)),
            'discharge_destination': np.random.choice(range(19, 99)),
            'primary_diagnosis': random.choice(list(conditions.keys())),
            'los_days': los,
            'readmission_flag': np.random.choice([0, 1], p=[0.85, 0.15])
        })

    # Generate QOF metrics
    qof_metrics = []
    qof_indicators = ['DM001', 'DM002', 'CHD001', 'HYP001', 'AST001', 'MH001', 'CAN001', 'COPD001', 'AF001', 'PAL001']
    for indicator in qof_indicators:
        qof_metrics.append({
            'indicator_code': indicator,
            'numerator': np.random.randint(100, 400),
            'denominator': np.random.randint(400, 500),
            'achievement_points': np.random.uniform(0, 100),
            'target_percentage': np.random.uniform(70, 95),
            'exception_reporting': np.random.uniform(0, 15)
        })

    return {
        'demographics': pd.DataFrame(demographics_data),
        'diagnoses': pd.DataFrame(diagnoses_records),
        'medications': pd.DataFrame(medications_records),
        'appointments': pd.DataFrame(appointments_records),
        'test_results': pd.DataFrame(test_results),
        'admissions': pd.DataFrame(admissions),
        'qof_metrics': pd.DataFrame(qof_metrics)
    }

def save_raw_data(data_dict, filepath):
    """Save raw data to Excel file with multiple sheets"""
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in data_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Raw data saved to {filepath}")

def create_human_friendly_report(data_dict, filepath):
    """Transform raw data into human-friendly report with charts"""

    wb = Workbook()

    # Create Executive Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"

    # Styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    subheader_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws_summary['A1'] = "NHS Integration Platform - Clinical Dashboard Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:H1')

    ws_summary['A3'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary['A4'] = f"Reporting Period: {(datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')} to {datetime.now().strftime('%Y-%m-%d')}"

    # Key Metrics
    ws_summary['A6'] = "KEY PERFORMANCE INDICATORS"
    ws_summary['A6'].font = header_font
    ws_summary['A6'].fill = header_fill
    ws_summary.merge_cells('A6:D6')

    total_patients = len(data_dict['demographics'])
    active_patients = len(data_dict['appointments']['patient_id'].unique())
    total_appointments = len(data_dict['appointments'])
    completed_appointments = len(data_dict['appointments'][data_dict['appointments']['status'] == 2])

    metrics = [
        ['Metric', 'Value', 'Target', 'Status'],
        ['Total Registered Patients', f"{total_patients:,}", '500', '✓'],
        ['Active Patients (with appointments)', f"{active_patients:,}", '400', '✓' if active_patients >= 400 else '✗'],
        ['Total Appointments', f"{total_appointments:,}", '1,800', '✓' if total_appointments >= 1800 else '✗'],
        ['Appointment Completion Rate', f"{(completed_appointments/total_appointments*100):.1f}%", '85%', '✓' if completed_appointments/total_appointments >= 0.85 else '✗'],
        ['Average Wait Time (days)', f"{data_dict['appointments']['wait_time_days'].mean():.1f}", '< 60', '✓' if data_dict['appointments']['wait_time_days'].mean() < 60 else '✗'],
        ['30-Day Readmission Rate', f"{(data_dict['admissions']['readmission_flag'].mean()*100):.1f}%", '< 20%', '✓' if data_dict['admissions']['readmission_flag'].mean() < 0.20 else '✗']
    ]

    for row_idx, row_data in enumerate(metrics, start=8):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 8:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = border

    # Create Patient Demographics sheet
    ws_demo = wb.create_sheet("Patient Demographics")

    # Transform demographics data
    conditions = generate_snomed_codes()
    demo_df = data_dict['demographics'].copy()
    demo_df['Gender'] = demo_df['gender_code'].map({1: 'Male', 2: 'Female', 9: 'Not Specified'})
    demo_df['Age'] = ((datetime.now() - pd.to_datetime(demo_df['dob'])).dt.days / 365.25).astype(int)
    demo_df['Age Group'] = pd.cut(demo_df['Age'], bins=[0, 18, 30, 50, 65, 100], labels=['0-17', '18-29', '30-49', '50-64', '65+'])

    # Add summary statistics to demographics sheet
    ws_demo['A1'] = "PATIENT DEMOGRAPHICS ANALYSIS"
    ws_demo['A1'].font = header_font
    ws_demo['A1'].fill = header_fill
    ws_demo.merge_cells('A1:F1')

    # Age distribution summary
    age_dist = demo_df['Age Group'].value_counts().sort_index()
    ws_demo['A3'] = "Age Distribution"
    ws_demo['A3'].font = subheader_font

    row = 4
    for age_group, count in age_dist.items():
        ws_demo[f'A{row}'] = age_group
        ws_demo[f'B{row}'] = count
        ws_demo[f'C{row}'] = f"{(count/len(demo_df)*100):.1f}%"
        row += 1

    # Gender distribution
    ws_demo['E3'] = "Gender Distribution"
    ws_demo['E3'].font = subheader_font

    gender_dist = demo_df['Gender'].value_counts()
    row = 4
    for gender, count in gender_dist.items():
        ws_demo[f'E{row}'] = gender
        ws_demo[f'F{row}'] = count
        ws_demo[f'G{row}'] = f"{(count/len(demo_df)*100):.1f}%"
        row += 1

    # Create Clinical Conditions sheet
    ws_clinical = wb.create_sheet("Clinical Conditions")

    diag_df = data_dict['diagnoses'].copy()
    diag_df['Condition'] = diag_df['snomed_code'].map(conditions)
    diag_df['Status'] = diag_df['status_code'].map({1: 'Active', 2: 'Resolved', 3: 'Inactive'})

    # Top conditions summary
    ws_clinical['A1'] = "TOP 10 CLINICAL CONDITIONS"
    ws_clinical['A1'].font = header_font
    ws_clinical['A1'].fill = header_fill
    ws_clinical.merge_cells('A1:D1')

    condition_counts = diag_df['Condition'].value_counts().head(10)

    ws_clinical['A3'] = 'Rank'
    ws_clinical['B3'] = 'Condition'
    ws_clinical['C3'] = 'Patient Count'
    ws_clinical['D3'] = 'Prevalence %'

    for col in ['A3', 'B3', 'C3', 'D3']:
        ws_clinical[col].font = Font(bold=True)
        ws_clinical[col].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for idx, (condition, count) in enumerate(condition_counts.items(), start=1):
        ws_clinical[f'A{idx+3}'] = idx
        ws_clinical[f'B{idx+3}'] = condition
        ws_clinical[f'C{idx+3}'] = count
        ws_clinical[f'D{idx+3}'] = f"{(count/len(diag_df['patient_id'].unique())*100):.1f}%"

    # Create Medications sheet
    ws_meds = wb.create_sheet("Medication Analysis")

    meds_dict = generate_medication_codes()
    meds_df = data_dict['medications'].copy()
    meds_df['Medication'] = meds_df['dm_d_code'].map(meds_dict)
    meds_df['Status'] = meds_df['status'].map({1: 'Active', 2: 'Discontinued', 3: 'On-hold'})

    ws_meds['A1'] = "MEDICATION PRESCRIBING PATTERNS"
    ws_meds['A1'].font = header_font
    ws_meds['A1'].fill = header_fill
    ws_meds.merge_cells('A1:E1')

    # Top prescribed medications
    med_counts = meds_df['Medication'].value_counts().head(10)

    ws_meds['A3'] = 'Rank'
    ws_meds['B3'] = 'Medication'
    ws_meds['C3'] = 'Prescriptions'
    ws_meds['D3'] = 'Avg Adherence'
    ws_meds['E3'] = 'Status'

    for col in ['A3', 'B3', 'C3', 'D3', 'E3']:
        ws_meds[col].font = Font(bold=True)
        ws_meds[col].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for idx, (med, count) in enumerate(med_counts.items(), start=1):
        ws_meds[f'A{idx+3}'] = idx
        ws_meds[f'B{idx+3}'] = med
        ws_meds[f'C{idx+3}'] = count
        avg_adherence = meds_df[meds_df['Medication'] == med]['adherence_score'].mean()
        ws_meds[f'D{idx+3}'] = f"{(avg_adherence*100):.1f}%"
        active_count = len(meds_df[(meds_df['Medication'] == med) & (meds_df['Status'] == 'Active')])
        ws_meds[f'E{idx+3}'] = f"{active_count}/{count} Active"

    # Create QOF Performance sheet
    ws_qof = wb.create_sheet("QOF Performance")

    qof_df = data_dict['qof_metrics'].copy()
    qof_df['Achievement Rate'] = (qof_df['numerator'] / qof_df['denominator'] * 100).round(1)
    qof_df['Target Met'] = qof_df['Achievement Rate'] >= qof_df['target_percentage']

    ws_qof['A1'] = "QUALITY OUTCOMES FRAMEWORK (QOF) PERFORMANCE"
    ws_qof['A1'].font = header_font
    ws_qof['A1'].fill = header_fill
    ws_qof.merge_cells('A1:F1')

    ws_qof['A3'] = 'Indicator'
    ws_qof['B3'] = 'Achievement'
    ws_qof['C3'] = 'Target'
    ws_qof['D3'] = 'Points'
    ws_qof['E3'] = 'Exception %'
    ws_qof['F3'] = 'Status'

    for col in ['A3', 'B3', 'C3', 'D3', 'E3', 'F3']:
        ws_qof[col].font = Font(bold=True)
        ws_qof[col].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for idx, row in qof_df.iterrows():
        ws_qof[f'A{idx+4}'] = row['indicator_code']
        ws_qof[f'B{idx+4}'] = f"{row['Achievement Rate']:.1f}%"
        ws_qof[f'C{idx+4}'] = f"{row['target_percentage']:.1f}%"
        ws_qof[f'D{idx+4}'] = f"{row['achievement_points']:.1f}"
        ws_qof[f'E{idx+4}'] = f"{row['exception_reporting']:.1f}%"
        ws_qof[f'F{idx+4}'] = '✓ Met' if row['Target Met'] else '✗ Not Met'

        if row['Target Met']:
            ws_qof[f'F{idx+4}'].font = Font(color="008000", bold=True)
        else:
            ws_qof[f'F{idx+4}'].font = Font(color="FF0000", bold=True)

    # Add Charts to existing sheets

    # Add pie chart to Patient Demographics sheet
    # Put chart data at the top, then chart below to avoid overlap
    demo_chart_data_row = 10
    ws_demo[f'J{demo_chart_data_row}'] = 'Chart Data'
    ws_demo[f'J{demo_chart_data_row}'].font = Font(bold=True)
    ws_demo[f'J{demo_chart_data_row+1}'] = 'Age Group'
    ws_demo[f'K{demo_chart_data_row+1}'] = 'Count'
    for idx, (age_group, count) in enumerate(age_dist.items(), start=1):
        ws_demo[f'J{demo_chart_data_row+1+idx}'] = age_group
        ws_demo[f'K{demo_chart_data_row+1+idx}'] = count

    pie = PieChart()
    labels = Reference(ws_demo, min_col=10, min_row=demo_chart_data_row+2, max_row=demo_chart_data_row+1+len(age_dist))
    data = Reference(ws_demo, min_col=11, min_row=demo_chart_data_row+1, max_row=demo_chart_data_row+1+len(age_dist))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Patient Age Distribution"
    pie.height = 10
    pie.width = 15
    ws_demo.add_chart(pie, "A20")

    # Add bar chart to Clinical Conditions sheet
    # Place chart data first, then chart below
    condition_chart_row = 3
    ws_clinical[f'F{condition_chart_row}'] = 'Chart Data'
    ws_clinical[f'F{condition_chart_row}'].font = Font(bold=True)
    ws_clinical[f'F{condition_chart_row+1}'] = 'Condition'
    ws_clinical[f'G{condition_chart_row+1}'] = 'Count'
    for idx, (condition, count) in enumerate(condition_counts.head(5).items(), start=1):
        ws_clinical[f'F{condition_chart_row+1+idx}'] = condition[:20]
        ws_clinical[f'G{condition_chart_row+1+idx}'] = count

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Top 5 Clinical Conditions"
    bar.y_axis.title = 'Number of Patients'
    bar.x_axis.title = 'Condition'

    data = Reference(ws_clinical, min_col=7, min_row=condition_chart_row+1, max_row=condition_chart_row+6)
    cats = Reference(ws_clinical, min_col=6, min_row=condition_chart_row+2, max_row=condition_chart_row+6)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height = 10
    bar.width = 15
    ws_clinical.add_chart(bar, "A16")

    # Add pie chart to Medications sheet
    med_counts = meds_df['Medication'].value_counts().head(6)
    med_chart_row = 3
    ws_meds[f'G{med_chart_row}'] = 'Chart Data'
    ws_meds[f'G{med_chart_row}'].font = Font(bold=True)
    ws_meds[f'G{med_chart_row+1}'] = 'Medication'
    ws_meds[f'H{med_chart_row+1}'] = 'Count'
    for idx, (med, count) in enumerate(med_counts.items(), start=1):
        ws_meds[f'G{med_chart_row+1+idx}'] = med[:25]
        ws_meds[f'H{med_chart_row+1+idx}'] = count

    pie2 = PieChart()
    labels2 = Reference(ws_meds, min_col=7, min_row=med_chart_row+2, max_row=med_chart_row+1+len(med_counts))
    data2 = Reference(ws_meds, min_col=8, min_row=med_chart_row+1, max_row=med_chart_row+1+len(med_counts))
    pie2.add_data(data2, titles_from_data=True)
    pie2.set_categories(labels2)
    pie2.title = "Top Prescribed Medications"
    pie2.height = 10
    pie2.width = 15
    ws_meds.add_chart(pie2, "A17")

    # Add bar chart to QOF Performance sheet
    qof_chart_row = 3
    ws_qof[f'H{qof_chart_row}'] = 'Chart Data'
    ws_qof[f'H{qof_chart_row}'].font = Font(bold=True)
    ws_qof[f'H{qof_chart_row+1}'] = 'Indicator'
    ws_qof[f'I{qof_chart_row+1}'] = 'Achievement'
    ws_qof[f'J{qof_chart_row+1}'] = 'Target'

    for idx, row in qof_df.head(5).iterrows():
        ws_qof[f'H{qof_chart_row+2+idx}'] = row['indicator_code']
        ws_qof[f'I{qof_chart_row+2+idx}'] = row['Achievement Rate']
        ws_qof[f'J{qof_chart_row+2+idx}'] = row['target_percentage']

    bar2 = BarChart()
    bar2.type = "col"
    bar2.style = 12
    bar2.title = "QOF Performance vs Targets"
    bar2.y_axis.title = 'Percentage'
    bar2.x_axis.title = 'Indicator'

    data1 = Reference(ws_qof, min_col=9, min_row=qof_chart_row+1, max_row=qof_chart_row+6)
    data2 = Reference(ws_qof, min_col=10, min_row=qof_chart_row+1, max_row=qof_chart_row+6)
    cats = Reference(ws_qof, min_col=8, min_row=qof_chart_row+2, max_row=qof_chart_row+6)

    bar2.add_data(data1, titles_from_data=True)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(cats)
    bar2.height = 10
    bar2.width = 15
    ws_qof.add_chart(bar2, "A16")

    # Add line chart to Executive Summary for appointment trends
    appt_df = data_dict['appointments'].copy()
    appt_df['appointment_date'] = pd.to_datetime(appt_df['appointment_date'])
    appt_df['month'] = appt_df['appointment_date'].dt.to_period('M')
    monthly_appts = appt_df.groupby('month').size().tail(6)

    summary_chart_row = 8
    ws_summary[f'F{summary_chart_row}'] = 'Chart Data'
    ws_summary[f'F{summary_chart_row}'].font = Font(bold=True)
    ws_summary[f'F{summary_chart_row+1}'] = 'Month'
    ws_summary[f'G{summary_chart_row+1}'] = 'Appointments'
    for idx, (month, count) in enumerate(monthly_appts.items(), start=1):
        ws_summary[f'F{summary_chart_row+1+idx}'] = str(month)[-7:]  # Show only YYYY-MM
        ws_summary[f'G{summary_chart_row+1+idx}'] = count

    line = LineChart()
    line.title = "6-Month Appointment Trends"
    line.style = 13
    line.y_axis.title = 'Appointments'
    line.x_axis.title = 'Month'

    data = Reference(ws_summary, min_col=7, min_row=summary_chart_row+1, max_row=summary_chart_row+1+len(monthly_appts))
    cats = Reference(ws_summary, min_col=6, min_row=summary_chart_row+2, max_row=summary_chart_row+1+len(monthly_appts))
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.height = 9
    line.width = 12
    ws_summary.add_chart(line, "A17")

    # Format column widths - simplified approach
    for sheet in wb.worksheets:
        # Set reasonable default widths for common columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            sheet.column_dimensions[col].width = 15

    # Save the workbook
    wb.save(filepath)
    print(f"Human-friendly report saved to {filepath}")

def main():
    print("NHS Integration Platform - Report Generator")
    print("=" * 50)

    # Generate raw data
    print("\n1. Generating raw NHS data...")
    raw_data = generate_raw_data()

    # Save raw data to input file
    input_file = os.path.join(public_folder, 'input-report.xlsx')
    save_raw_data(raw_data, input_file)

    # Create human-friendly report
    print("\n2. Creating human-friendly report with visualizations...")
    output_file = os.path.join(public_folder, 'sample-report.xlsx')
    create_human_friendly_report(raw_data, output_file)

    print("\n" + "=" * 50)
    print("Report generation complete!")
    print(f"\nFiles created:")
    print(f"1. Input data (raw): {input_file}")
    print(f"2. Output report (human-friendly): {output_file}")

    # Display summary statistics
    print("\n" + "=" * 50)
    print("DATA SUMMARY:")
    print(f"- Total patients: {len(raw_data['demographics'])}")
    print(f"- Total diagnoses: {len(raw_data['diagnoses'])}")
    print(f"- Total medications: {len(raw_data['medications'])}")
    print(f"- Total appointments: {len(raw_data['appointments'])}")
    print(f"- Total test results: {len(raw_data['test_results'])}")
    print(f"- Total admissions: {len(raw_data['admissions'])}")
    print(f"- QOF indicators tracked: {len(raw_data['qof_metrics'])}")

if __name__ == "__main__":
    main()